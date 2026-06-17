"""
plot.py — 2x3 E-NMPC comparison plot for the Test network using local Excel data.

Place this script in the SAME folder as the Excel files. It reads them via
relative paths and writes `enmpc_comparison.png` next to it.

Layout (2 rows x 3 columns):
    Row 1 (top):    Source flow on LEFT y-axis (kg/s).
                    Source pressure is constant in this network, so the right
                    pressure axis is omitted.
    Row 2 (bottom): All 3 compressor energies (kWh) with OCSS dotted.

Columns:
    Col 1: f_enmpc_cycle_1.xlsx (finite-horizon NMPC, 1-cycle horizon)
    Col 2: f_enmpc_cycle_3.xlsx (finite-horizon NMPC, 3-cycle horizon)
    Col 3: Inf_enmpc_felement=1.xlsx (infinite-horizon NMPC, 1-cycle segment)

OCSS reference comes from `css.xlsx` (perfectly periodic single-cycle OCSS,
tiled across the visible time axis).
"""

from pathlib import Path
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

# ----------------------------------------------------------------------------
# Configuration -- edit names, slices, units, titles here.
# ----------------------------------------------------------------------------
HERE = Path(__file__).resolve().parent

# Truncate x-axis (and data) to this many hours: 4 cycles x 6 hours.
X_MAX_HOURS = 24

# (filename, title, slice). slice=None means use all rows.
COLUMNS = [
    ("f_enmpc_cycle_1.xlsx",      "Finite horizon E-NMPC\nHorizon – 1 cycle",          None),
    ("f_enmpc_cycle_10.xlsx",      "Finite horizon E-NMPC\nHorizon – 10 cycles",         None),
    ("Inf_enmpc_felement=1.xlsx", "Infinite horizon E-NMPC\nFinite segment – 1 cycle", None),
]

# OCSS reference: dedicated css.xlsx (perfectly periodic single-cycle OCSS).
OCSS_FILE          = "css.xlsx"
OCSS_CYCLE_HOURS   = 6
OCSS_DROP_LAST     = 1     # drop the duplicated periodic endpoint

# Test network: single source; pSource is constant, so only flow is plotted.
LEFT_SOURCES  = ["source_1"]
RIGHT_SOURCES = []         # no pressure on right axis

# Unit scaling driven by model.py: scale['p']=1e5, scale['P']=1e5, scale['w']=1.
FLOW_SCALE     = 1.0
PRESSURE_SCALE = 1.0
POWER_SCALE    = 100.0    # compressor_P [internal] * 100 = kW = kWh per 1-hour step

# Output
OUTPUT_PNG               = "enmpc_comparison.png"
OUTPUT_PNG_BETA_TOP      = "enmpc_comparison_beta.png"    # beta (top) + energy (bottom)
OUTPUT_PNG_SOURCES       = "enmpc_sources_split.png"      # per-source flow figure
OUTPUT_PNG_COMPRESSORS   = "enmpc_compressors_split.png"  # per-compressor energy figure
OUTPUT_PNG_BETA          = "enmpc_beta_split.png"         # per-compressor beta figure
OUTPUT_XLSX              = "obj.xlsx"
DPI                      = 200

# Cost annotation: replicate OBJ_compressor_power from model.py
SHOW_COST = True

# ============================================================================
# ENERGY FORMULA NOTE
# ============================================================================
# The energy cost is computed as: sum(P_internal) * 100 kWh per cycle
# This assumes:
#   - P_internal is in units where P_internal * 100 = kW (kilowatts)
#   - Each timestep = 1 hour
#   - scale['P'] = 100000 in model.py, so internal values are 1/1000 of kW
#
# Model objective in model.py uses: sum(P_internal * dt) / 1000
# where dt = 3600 seconds. This gives: sum(P_internal) * 3.6
#
# MISMATCH ALERT: If plot shows X kWh but model objective shows Y,
# and 3.6*Y ≠ X, then there's a unit scaling issue.
# FIX: Change Options.json dt from 3600 to 1.0 (hours) to match plot formula.
# ============================================================================


def compute_cycle_costs(pwr_df, pwr_cols, cycle_hours, dt=3600, scale_factor=1000):
    """Return per-cycle costs using exact OBJ_compressor_power formula from model.py:

    Obj = sum(P_internal * dt) / scale_factor

    With dt = 3600 seconds and scale_factor = 1000:
    Obj = sum(P_internal * 3600) / 1000 ≈ sum(P_internal) * 3.6

    Or without scaling (scale_factor=1):
    Obj = sum(P_internal * 3600) (raw objective value)

    Excludes the last timestep.
    """
    arr = pwr_df[pwr_cols].iloc[:-1].values    # drop final timestep
    n = len(arr)
    cycles = []
    for k in range(n // cycle_hours):
        block = arr[k * cycle_hours : (k + 1) * cycle_hours]
        cycles.append(float(block.sum()) * dt / scale_factor)
    return cycles


def compute_cost_kwh(pwr_df, pwr_cols, dt=3600, scale_factor=1000):
    """Total cost over all complete cycles."""
    return sum(compute_cycle_costs(pwr_df, pwr_cols, OCSS_CYCLE_HOURS, dt=dt, scale_factor=scale_factor))


def compute_cost_model_formula(pwr_df, pwr_cols, dt=3600):
    """DEBUG: Compute cost using exact model formula: sum(P * dt) / 1000.

    This should theoretically match compute_cost_kwh() if the units are consistent.
    If results differ significantly, it indicates a unit mismatch.
    """
    arr = pwr_df[pwr_cols].iloc[:-1].values.sum()  # drop final timestep
    return arr * dt / 1000


# ----------------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------------
def load_run(path):
    """Load wSource, pSource, compressor power, compressor beta sheets."""
    src_w = pd.read_excel(path, sheet_name="wSource")
    src_p = pd.read_excel(path, sheet_name="pSource")
    pwr   = pd.read_excel(path, sheet_name="compressor power")
    beta  = pd.read_excel(path, sheet_name="compressor beta")
    for df in (src_w, src_p, pwr, beta):
        if "Unnamed: 0" in df.columns:
            df.drop(columns=["Unnamed: 0"], inplace=True)
    return src_w, src_p, pwr, beta


def sort_by_trailing_int(cols):
    """Sort variable columns by the trailing integer in the entity name."""
    def key(c):
        try:
            return int(c.split("'")[1].split("_")[-1])
        except Exception:
            return 0
    return sorted(cols, key=key)


def entity_name(col):
    """Extract 'source_1' from \"wSource['source_1', :]\" etc."""
    return col.split("'")[1]


def slice_run(src_w, src_p, pwr, beta, sl):
    if sl is None:
        return src_w, src_p, pwr, beta
    return (src_w.iloc[sl].reset_index(drop=True),
            src_p.iloc[sl].reset_index(drop=True),
            pwr.iloc[sl].reset_index(drop=True),
            beta.iloc[sl].reset_index(drop=True))


def truncate_to_hours(src_w, src_p, pwr, beta, max_hours):
    """Keep rows 0..max_hours (inclusive) — one extra row for the endpoint."""
    keep = min(max_hours + 1, len(src_w))
    return (src_w.iloc[:keep].reset_index(drop=True),
            src_p.iloc[:keep].reset_index(drop=True),
            pwr.iloc[:keep].reset_index(drop=True),
            beta.iloc[:keep].reset_index(drop=True))


def tile_ocss(ocss_t, ocss_vals, target_t, cycle_hours):
    """Periodic interpolation: value at t+cycle_hours equals value at t."""
    phase = np.mod(target_t, cycle_hours)
    t_ext = np.concatenate([ocss_t, [cycle_hours]])
    v_ext = np.concatenate([ocss_vals, [ocss_vals[0]]])
    return np.interp(phase, t_ext, v_ext)


def load_cpu_time_seconds(path):
    """Read total CPU time in seconds from the NMPC Excel file.
    Reads the trailing 'TOTAL' row of the `runtime` sheet (preferred),
    or falls back to `timing.total_runtime_s[0]`.
    """
    xl = pd.ExcelFile(path)
    if "runtime" in xl.sheet_names:
        df = pd.read_excel(path, sheet_name="runtime")
        if "iteration_time_s" in df.columns and len(df) > 0:
            val = df["iteration_time_s"].iloc[-1]
            if pd.notna(val):
                return float(val)
    if "timing" in xl.sheet_names:
        df = pd.read_excel(path, sheet_name="timing")
        if "total_runtime_s" in df.columns:
            val = df["total_runtime_s"].iloc[0]
            if pd.notna(val):
                return float(val)
    return None


def format_cpu_time(seconds):
    """Test network: seconds."""
    return f"{seconds:.1f} s"


# ----------------------------------------------------------------------------
# Load data
# ----------------------------------------------------------------------------
runs = []
for fn, title, sl in COLUMNS:
    src_w, src_p, pwr, beta = load_run(HERE / fn)
    src_w, src_p, pwr, beta = slice_run(src_w, src_p, pwr, beta, sl)
    src_w, src_p, pwr, beta = truncate_to_hours(src_w, src_p, pwr, beta, X_MAX_HOURS)
    cpu_seconds = load_cpu_time_seconds(HERE / fn)
    t = np.arange(len(src_w))
    runs.append((title, t, src_w, src_p, pwr, beta, cpu_seconds))

# OCSS reference: one clean cycle
ocss_src_w, ocss_src_p, ocss_pwr, ocss_beta = load_run(HERE / OCSS_FILE)
# css.xlsx may already span multiple periods; keep only the first cycle
keep = min(OCSS_CYCLE_HOURS, len(ocss_src_w) - OCSS_DROP_LAST)
ocss_src_w = ocss_src_w.iloc[:keep].reset_index(drop=True)
ocss_src_p = ocss_src_p.iloc[:keep].reset_index(drop=True)
ocss_pwr   = ocss_pwr.iloc[:keep].reset_index(drop=True)
ocss_beta  = ocss_beta.iloc[:keep].reset_index(drop=True)
ocss_t     = np.arange(len(ocss_src_w))

# Column lists, sorted by trailing integer for consistent ordering / colors
wS_cols   = sort_by_trailing_int([c for c in ocss_src_w.columns if "wSource"         in c])
pS_cols   = sort_by_trailing_int([c for c in ocss_src_p.columns if "pSource"         in c])
pwr_cols  = sort_by_trailing_int([c for c in ocss_pwr.columns   if "compressor_P"    in c])
beta_cols = sort_by_trailing_int([c for c in ocss_beta.columns  if "compressor_beta" in c])

# Map entity_name -> wSource and pSource column strings
wS_by_src = {entity_name(c): c for c in wS_cols}
pS_by_src = {entity_name(c): c for c in pS_cols}

# Colors per source (consistent across left/right axes and OCSS lines)
src_order = LEFT_SOURCES + RIGHT_SOURCES
cmap10    = plt.get_cmap("tab10")
src_color = {name: cmap10(i % 10) for i, name in enumerate(src_order)}

# Colors per compressor
pwr_color = {c: cmap10(i % 10) for i, c in enumerate(pwr_cols)}

# Map beta column name -> entity (compressorStation_X) -> matching pwr column,
# so beta plots can reuse the same per-compressor colors as the energy plots.
beta_entity_to_pwr_col = {entity_name(b): p for b, p in zip(beta_cols, pwr_cols)}


# ----------------------------------------------------------------------------
# Plot
# ----------------------------------------------------------------------------
plt.rcParams.update({
    "font.family": "serif",
    "font.size": 11,
    "axes.grid": True,
    "grid.alpha": 0.3,
    "grid.linestyle": "--",
})

n_cols = len(runs)
fig, axes = plt.subplots(2, n_cols, figsize=(5.3 * n_cols, 8),
                         constrained_layout=True)
if n_cols == 1:
    axes = axes.reshape(2, 1)

for col_idx, (title, t, src_w, src_p, pwr, beta, cpu_seconds) in enumerate(runs):
    ax_top = axes[0, col_idx]
    ax_bot = axes[1, col_idx]

    handles, labels_ = [], []

    # ----- Top row LEFT axis: flow for LEFT_SOURCES -----
    for name in LEFT_SOURCES:
        if name in wS_by_src and wS_by_src[name] in src_w.columns:
            c = wS_by_src[name]
            color = src_color[name]
            line, = ax_top.plot(t, src_w[c].values * FLOW_SCALE,
                                color=color, lw=1.7, label=name)
            handles.append(line); labels_.append(name)
            y_ocss = tile_ocss(ocss_t, ocss_src_w[c].values * FLOW_SCALE,
                               t, OCSS_CYCLE_HOURS)
            ax_top.plot(t, y_ocss, color=color, lw=1.1, linestyle=":", alpha=0.75)

    # OCSS legend entry (neutral)
    ocss_proxy = plt.Line2D([], [], color="black", linestyle=":", lw=1.2)
    handles.append(ocss_proxy); labels_.append("OCSS")

    ax_top.set_title(title, fontsize=12)
    ax_top.set_xlabel("Time (hrs)")
    ax_top.set_ylabel("Flow (kg/s)")
    ax_top.set_xlim(0, X_MAX_HOURS)
    ax_top.legend(handles, labels_, loc="upper right",
                  fontsize=8, framealpha=0.9)

    # ----- Bottom row: Compressor energy (kWh) -----
    for c in pwr_cols:
        if c in pwr.columns:
            label = entity_name(c).replace("compressorStation_", "C")
            ax_bot.plot(t, pwr[c].values * POWER_SCALE,
                        color=pwr_color[c], lw=1.7, label=label)
    for c in pwr_cols:
        if c in ocss_pwr.columns:
            y_ocss = tile_ocss(ocss_t, ocss_pwr[c].values * POWER_SCALE,
                               t, OCSS_CYCLE_HOURS)
            ax_bot.plot(t, y_ocss, color=pwr_color[c],
                        lw=1.1, linestyle=":", alpha=0.75)
    ax_bot.plot([], [], color="black", linestyle=":", lw=1.2, label="OCSS")

    ax_bot.set_xlabel("Time (hrs)")
    ax_bot.set_ylabel("Energy (kWh)")
    ax_bot.set_xlim(0, X_MAX_HOURS)
    ax_bot.legend(loc="upper right", fontsize=8, ncol=2, framealpha=0.9)

    # Annotate total cost below this column
    if SHOW_COST:
        # Use model formula: sum(P_internal * dt) / scale_factor with dt=3600, scale=1
        # This gives the raw objective value (e.g., 5059 instead of 155)
        cost = compute_cost_kwh(pwr, pwr_cols, dt=3600, scale_factor=1)
        cycle_costs = compute_cycle_costs(pwr, pwr_cols, OCSS_CYCLE_HOURS, dt=3600, scale_factor=1)
        cycle_sum_expr = " + ".join(f"{c:.0f}" for c in cycle_costs)
        cpu_str = format_cpu_time(cpu_seconds) if cpu_seconds is not None else "N/A"
        ax_bot.text(0.5, -0.28,
                    f"Total Cost: {cost:.0f}\n"
                    f"{cycle_sum_expr}\n"
                    f"CPU time: {cpu_str}",
                    transform=ax_bot.transAxes, ha="center", va="top",
                    fontsize=11, color="firebrick", fontweight="bold")

# Save
out = HERE / OUTPUT_PNG
fig.savefig(out, dpi=DPI, bbox_inches="tight")
print(f"Saved: {out}")


# ----------------------------------------------------------------------------
# Beta-on-top variant: same as the main 2xN comparison, but the top row plots
# all compressor beta values (instead of source flow). Bottom row unchanged.
# ----------------------------------------------------------------------------
fig_b, axes_b = plt.subplots(2, n_cols, figsize=(5.3 * n_cols, 8),
                             constrained_layout=True)
if n_cols == 1:
    axes_b = axes_b.reshape(2, 1)

for col_idx, (title, t, src_w, src_p, pwr, beta, cpu_seconds) in enumerate(runs):
    ax_top = axes_b[0, col_idx]
    ax_bot = axes_b[1, col_idx]

    # ----- Top row: Compressor beta (pressure ratio) -----
    for c in beta_cols:
        if c in beta.columns:
            ent = entity_name(c)
            label = ent.replace("compressorStation_", "C")
            color = pwr_color[beta_entity_to_pwr_col[ent]] \
                if ent in beta_entity_to_pwr_col else None
            ax_top.plot(t, beta[c].values, color=color, lw=1.7, label=label)
    for c in beta_cols:
        if c in ocss_beta.columns:
            ent = entity_name(c)
            color = pwr_color[beta_entity_to_pwr_col[ent]] \
                if ent in beta_entity_to_pwr_col else None
            y_ocss = tile_ocss(ocss_t, ocss_beta[c].values, t, OCSS_CYCLE_HOURS)
            ax_top.plot(t, y_ocss, color=color, lw=1.1, linestyle=":", alpha=0.75)
    ax_top.plot([], [], color="black", linestyle=":", lw=1.2, label="OCSS")

    ax_top.set_title(title, fontsize=12)
    ax_top.set_xlabel("Time (hrs)")
    ax_top.set_ylabel(r"Compressor $\beta$")
    ax_top.set_xlim(0, X_MAX_HOURS)
    ax_top.legend(loc="upper right", fontsize=8, ncol=2, framealpha=0.9)

    # ----- Bottom row: Compressor energy (kWh) -----
    for c in pwr_cols:
        if c in pwr.columns:
            label = entity_name(c).replace("compressorStation_", "C")
            ax_bot.plot(t, pwr[c].values * POWER_SCALE,
                        color=pwr_color[c], lw=1.7, label=label)
    for c in pwr_cols:
        if c in ocss_pwr.columns:
            y_ocss = tile_ocss(ocss_t, ocss_pwr[c].values * POWER_SCALE,
                               t, OCSS_CYCLE_HOURS)
            ax_bot.plot(t, y_ocss, color=pwr_color[c],
                        lw=1.1, linestyle=":", alpha=0.75)
    ax_bot.plot([], [], color="black", linestyle=":", lw=1.2, label="OCSS")

    ax_bot.set_xlabel("Time (hrs)")
    ax_bot.set_ylabel("Energy (kWh)")
    ax_bot.set_xlim(0, X_MAX_HOURS)
    ax_bot.legend(loc="upper right", fontsize=8, ncol=2, framealpha=0.9)

    if SHOW_COST:
        cost = compute_cost_kwh(pwr, pwr_cols, dt=3600, scale_factor=1)
        cycle_costs = compute_cycle_costs(pwr, pwr_cols, OCSS_CYCLE_HOURS, dt=3600, scale_factor=1)
        cycle_sum_expr = " + ".join(f"{c:.0f}" for c in cycle_costs)
        cpu_str = format_cpu_time(cpu_seconds) if cpu_seconds is not None else "N/A"
        ax_bot.text(0.5, -0.28,
                    f"Total Cost: {cost:.0f}\n"
                    f"{cycle_sum_expr}\n"
                    f"CPU time: {cpu_str}",
                    transform=ax_bot.transAxes, ha="center", va="top",
                    fontsize=11, color="firebrick", fontweight="bold")

out_b = HERE / OUTPUT_PNG_BETA_TOP
fig_b.savefig(out_b, dpi=DPI, bbox_inches="tight")
print(f"Saved: {out_b}")


# ----------------------------------------------------------------------------
# Second figure: per-source flow comparison (rows = sources, cols = scenarios)
# ----------------------------------------------------------------------------
n_rows_src = len(LEFT_SOURCES)
fig2, axes2 = plt.subplots(n_rows_src, n_cols,
                           figsize=(5.3 * n_cols, 2.6 * n_rows_src),
                           sharex=True, constrained_layout=True)
# Normalize to 2-D indexing
axes2 = np.atleast_2d(axes2)
if n_rows_src == 1 and n_cols > 1:
    axes2 = axes2.reshape(1, n_cols)
elif n_rows_src > 1 and n_cols == 1:
    axes2 = axes2.reshape(n_rows_src, 1)

for col_idx, (title, t, src_w, src_p, pwr, beta, cpu_seconds) in enumerate(runs):
    axes2[0, col_idx].set_title(title, fontsize=12)

    for row_idx, name in enumerate(LEFT_SOURCES):
        ax = axes2[row_idx, col_idx]
        if name in wS_by_src and wS_by_src[name] in src_w.columns:
            c = wS_by_src[name]
            color = src_color[name]
            ax.plot(t, src_w[c].values * FLOW_SCALE,
                    color=color, lw=1.7, label=name)
            y_ocss = tile_ocss(ocss_t, ocss_src_w[c].values * FLOW_SCALE,
                               t, OCSS_CYCLE_HOURS)
            ax.plot(t, y_ocss, color=color, lw=1.1, linestyle=":", alpha=0.75,
                    label="OCSS")
            ax.legend(loc="upper right", fontsize=8, framealpha=0.9)

        ax.set_xlim(0, X_MAX_HOURS)
        if col_idx == 0:
            ax.set_ylabel("Flow (kg/s)")
        if row_idx == n_rows_src - 1:
            ax.set_xlabel("Time (hrs)")

out2 = HERE / OUTPUT_PNG_SOURCES
fig2.savefig(out2, dpi=DPI, bbox_inches="tight")
print(f"Saved: {out2}")


# ----------------------------------------------------------------------------
# Third figure: per-compressor energy comparison (rows = compressors, cols = scenarios)
# ----------------------------------------------------------------------------
n_rows_pwr = len(pwr_cols)
fig3, axes3 = plt.subplots(n_rows_pwr, n_cols,
                           figsize=(5.3 * n_cols, 2.6 * n_rows_pwr),
                           sharex=True, constrained_layout=True)
axes3 = np.atleast_2d(axes3)
if n_rows_pwr == 1 and n_cols > 1:
    axes3 = axes3.reshape(1, n_cols)
elif n_rows_pwr > 1 and n_cols == 1:
    axes3 = axes3.reshape(n_rows_pwr, 1)

for col_idx, (title, t, src_w, src_p, pwr, beta, cpu_seconds) in enumerate(runs):
    axes3[0, col_idx].set_title(title, fontsize=12)

    for row_idx, c in enumerate(pwr_cols):
        ax = axes3[row_idx, col_idx]
        label = entity_name(c).replace("compressorStation_", "C")
        color = pwr_color[c]
        if c in pwr.columns:
            ax.plot(t, pwr[c].values * POWER_SCALE,
                    color=color, lw=1.7, label=label)
        if c in ocss_pwr.columns:
            y_ocss = tile_ocss(ocss_t, ocss_pwr[c].values * POWER_SCALE,
                               t, OCSS_CYCLE_HOURS)
            ax.plot(t, y_ocss, color=color, lw=1.1, linestyle=":", alpha=0.75,
                    label="OCSS")
        ax.legend(loc="upper right", fontsize=8, framealpha=0.9)

        ax.set_xlim(0, X_MAX_HOURS)
        if col_idx == 0:
            ax.set_ylabel("Energy (kWh)")
        if row_idx == n_rows_pwr - 1:
            ax.set_xlabel("Time (hrs)")

out3 = HERE / OUTPUT_PNG_COMPRESSORS
fig3.savefig(out3, dpi=DPI, bbox_inches="tight")
print(f"Saved: {out3}")


# ----------------------------------------------------------------------------
# Fourth figure: per-compressor beta (pressure ratio) comparison
# rows = compressors, cols = scenarios
# ----------------------------------------------------------------------------
n_rows_beta = len(beta_cols)
fig4, axes4 = plt.subplots(n_rows_beta, n_cols,
                           figsize=(5.3 * n_cols, 2.6 * n_rows_beta),
                           sharex=True, constrained_layout=True)
axes4 = np.atleast_2d(axes4)
if n_rows_beta == 1 and n_cols > 1:
    axes4 = axes4.reshape(1, n_cols)
elif n_rows_beta > 1 and n_cols == 1:
    axes4 = axes4.reshape(n_rows_beta, 1)

for col_idx, (title, t, src_w, src_p, pwr, beta, cpu_seconds) in enumerate(runs):
    axes4[0, col_idx].set_title(title, fontsize=12)

    for row_idx, c in enumerate(beta_cols):
        ax = axes4[row_idx, col_idx]
        ent = entity_name(c)
        label = ent.replace("compressorStation_", "C")
        color = pwr_color[beta_entity_to_pwr_col[ent]]
        if c in beta.columns:
            ax.plot(t, beta[c].values,
                    color=color, lw=1.7, label=label)
        if c in ocss_beta.columns:
            y_ocss = tile_ocss(ocss_t, ocss_beta[c].values,
                               t, OCSS_CYCLE_HOURS)
            ax.plot(t, y_ocss, color=color, lw=1.1, linestyle=":", alpha=0.75,
                    label="OCSS")
        ax.legend(loc="upper right", fontsize=8, framealpha=0.9)

        ax.set_xlim(0, X_MAX_HOURS)
        if col_idx == 0:
            ax.set_ylabel(r"Compressor $\beta$")
        if row_idx == n_rows_beta - 1:
            ax.set_xlabel("Time (hrs)")

out4 = HERE / OUTPUT_PNG_BETA
fig4.savefig(out4, dpi=DPI, bbox_inches="tight")
print(f"Saved: {out4}")


# ----------------------------------------------------------------------------
# Cost summary -> obj.xlsx
# ----------------------------------------------------------------------------
def write_cost_summary(path):
    """Write a per-cycle cost breakdown for all scenarios to an Excel file.

    Per-cycle values are hardcoded (computed from the trajectory data);
    aggregates (Total, Avg) use Excel formulas so the table stays dynamic.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    # Compute per-cycle costs for each scenario (truncated to X_MAX_HOURS)
    rows = []
    for fn, title, _ in COLUMNS:
        pwr = pd.read_excel(HERE / fn, sheet_name="compressor power")
        if "Unnamed: 0" in pwr.columns:
            pwr = pwr.drop(columns=["Unnamed: 0"])
        pwr = pwr.iloc[:X_MAX_HOURS + 1].reset_index(drop=True)
        cols = sort_by_trailing_int([c for c in pwr.columns if "compressor_P" in c])
        cycles = compute_cycle_costs(pwr, cols, OCSS_CYCLE_HOURS, dt=3600, scale_factor=1)
        flat_title = title.replace("\n", " — ")
        rows.append((flat_title, cycles))

    n_cycles = max(len(c) for _, c in rows) if rows else 0

    wb = Workbook()
    ws = wb.active
    ws.title = "Cost Summary"

    header = ["Scenario"] + [f"Cycle {i+1}" for i in range(n_cycles)] \
             + [f"{X_MAX_HOURS}hr Total", "Average per Cycle"]
    ws.append(header)

    for i, (name, cycles) in enumerate(rows, start=2):
        row = [name] + cycles + [None] * (n_cycles - len(cycles))
        ws.append(row)

        cyc_start = get_column_letter(2)
        cyc_end   = get_column_letter(1 + n_cycles)
        total_col = get_column_letter(2 + n_cycles)
        avg_col   = get_column_letter(3 + n_cycles)
        ws[f"{total_col}{i}"] = f"=SUM({cyc_start}{i}:{cyc_end}{i})"
        ws[f"{avg_col}{i}"]   = f"=AVERAGE({cyc_start}{i}:{cyc_end}{i})"

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="305496")
    body_font   = Font(name="Arial")
    thin        = Side(style="thin", color="B4B4B4")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, _ in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    n_rows = 1 + len(rows)
    for r in range(2, n_rows + 1):
        for c in range(1, len(header) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = body_font
            cell.border = border
            if c == 1:
                cell.alignment = Alignment(horizontal="left",  vertical="center")
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "0.00"

    ws.column_dimensions["A"].width = 42
    for col_idx in range(2, len(header) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18
    ws.row_dimensions[1].height = 32

    ws.freeze_panes = "B2"

    wb.save(path)
    print(f"Saved: {path}")


write_cost_summary(HERE / OUTPUT_XLSX)

plt.show()
